#!/usr/bin/env python3
"""
SMS Script for Wedding - Send seat numbers and photo upload links to guests

This script reads the RSVP Excel file and sends personalized SMS messages to all guests
with their seat number and a link to upload wedding photos.

Requirements:
    pip install twilio openpyxl

Twilio Setup:
    1. Sign up at https://www.twilio.com/
    2. Get your Account SID and Auth Token from the console
    3. Get a Twilio phone number
    4. Update the credentials below

Usage:
    python send_sms_notifications.py <path_to_excel_file>

    Example:
    python send_sms_notifications.py wedding_rsvps_20260904.xlsx
"""

import sys
import os
from openpyxl import load_workbook
from twilio.rest import Client

# Twilio credentials - REPLACE THESE WITH YOUR ACTUAL CREDENTIALS
TWILIO_ACCOUNT_SID = 'your_account_sid_here'
TWILIO_AUTH_TOKEN = 'your_auth_token_here'
TWILIO_PHONE_NUMBER = '+1234567890'  # Your Twilio phone number

# Website URL for photo uploads
PHOTO_UPLOAD_URL = 'https://loevliedl.com/wedding/photos/upload/'

# Test mode - set to True to print messages instead of sending
TEST_MODE = True


def format_phone_number(phone):
    """Format phone number to E.164 format (+1XXXXXXXXXX)"""
    # Remove all non-digit characters
    digits = ''.join(filter(str.isdigit, str(phone)))

    # Add +1 if not present (assuming US numbers)
    if not digits.startswith('1') and len(digits) == 10:
        digits = '1' + digits

    return '+' + digits


def send_sms(client, to_phone, message):
    """Send SMS message via Twilio"""
    try:
        message = client.messages.create(
            body=message,
            from_=TWILIO_PHONE_NUMBER,
            to=to_phone
        )
        return True, message.sid
    except Exception as e:
        return False, str(e)


def main():
    if len(sys.argv) < 2:
        print("Usage: python send_sms_notifications.py <path_to_excel_file>")
        print("Example: python send_sms_notifications.py wedding_rsvps_20260904.xlsx")
        sys.exit(1)

    excel_file = sys.argv[1]

    if not os.path.exists(excel_file):
        print(f"Error: File '{excel_file}' not found!")
        sys.exit(1)

    print(f"Reading Excel file: {excel_file}")
    print(f"Test mode: {TEST_MODE}")
    print("-" * 60)

    # Initialize Twilio client (only if not in test mode)
    client = None
    if not TEST_MODE:
        if TWILIO_ACCOUNT_SID == 'your_account_sid_here':
            print("ERROR: Please update Twilio credentials in the script!")
            sys.exit(1)
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    # Load Excel file
    wb = load_workbook(excel_file)
    ws = wb.active

    # Expected headers: First Name, Last Name, Contact Phone, Email, Is Primary Contact,
    #                   Attending, Dietary Restrictions, Song Request, Message, Submitted At, Seat Number

    sent_count = 0
    error_count = 0
    skipped_count = 0

    # Skip header row
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        first_name = row[0]
        last_name = row[1]
        contact_phone = row[2]
        attending = row[5]
        seat_number = row[10]  # Column K (Seat Number)

        # Skip if not attending or no seat number assigned
        if not attending or attending.lower() != 'accept':
            print(f"Row {row_num}: Skipping {first_name} {last_name} - Not attending")
            skipped_count += 1
            continue

        if not seat_number:
            print(f"Row {row_num}: Skipping {first_name} {last_name} - No seat number assigned")
            skipped_count += 1
            continue

        if not contact_phone:
            print(f"Row {row_num}: Skipping {first_name} {last_name} - No phone number")
            skipped_count += 1
            continue

        # Format phone number
        try:
            formatted_phone = format_phone_number(contact_phone)
        except Exception as e:
            print(f"Row {row_num}: Error formatting phone for {first_name} {last_name}: {e}")
            error_count += 1
            continue

        # Create personalized message
        message = f"""Hi {first_name}! 🐧💕🐧

Thank you for celebrating with Dennis & Caitlin!

Your seat number is: {seat_number}

Please share your photos from the wedding:
{PHOTO_UPLOAD_URL}

We can't wait to see the memories you captured!

- Dennis & Caitlin"""

        # Send or print message
        if TEST_MODE:
            print(f"\nRow {row_num}: Would send to {first_name} {last_name} ({formatted_phone}):")
            print(f"Seat: {seat_number}")
            print(f"Message:\n{message}")
            print("-" * 60)
            sent_count += 1
        else:
            print(f"Sending to {first_name} {last_name} ({formatted_phone})...", end=" ")
            success, result = send_sms(client, formatted_phone, message)
            if success:
                print(f"✓ Sent (SID: {result})")
                sent_count += 1
            else:
                print(f"✗ Failed: {result}")
                error_count += 1

    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total messages sent: {sent_count}")
    print(f"Errors: {error_count}")
    print(f"Skipped: {skipped_count}")

    if TEST_MODE:
        print("\n⚠️  TEST MODE - No messages were actually sent!")
        print("Set TEST_MODE = False to send real messages.")


if __name__ == '__main__':
    main()
