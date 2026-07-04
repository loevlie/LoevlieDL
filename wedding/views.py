from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from .models import Location, WeddingPartyMember, RSVP, Guest, PhotoUpload
from .forms import RSVPForm


def home(request):
    """Wedding landing page with hero section"""
    context = {
        'bride_name': 'Caitlin Morrow',
        'groom_name': 'Dennis Loevlie',
        'wedding_date': 'September 5th, 2026',
        'wedding_location': 'Pittsburgh, PA',
    }
    return render(request, 'wedding/home.html', context)


def our_story(request):
    """Page displaying the couple's story"""
    story_text = """We met in undergrad as classmates in the same major, stayed friends,
    and finally started dating senior year. After graduation, we moved to Pittsburgh and
    began building a life we love—exploring neighborhoods, taking weekend trips, and finding
    our spots around the city. At home, we cook together and unwind with romantic comedies.
    From lecture halls to city streets, we're still choosing each other—and can't wait to
    celebrate with you."""

    context = {
        'story_text': story_text,
    }
    return render(request, 'wedding/our_story.html', context)


def wedding_party(request):
    """Display wedding party members with photos and carousel"""
    party_members = WeddingPartyMember.objects.filter(is_active=True)

    context = {
        'party_members': party_members,
        'bride_side': party_members.filter(side='bride'),
        'groom_side': party_members.filter(side='groom'),
    }
    return render(request, 'wedding/wedding_party.html', context)


def party_photos_api(request):
    """API endpoint to list all photos from Cloudinary"""
    import cloudinary.api

    try:
        # Get all images from Cloudinary wedding/party folder
        result = cloudinary.api.resources(
            type="upload",
            prefix="wedding/party/",
            max_results=500
        )

        photos = []
        seen_hashes = {}  # Track by etag/hash to find true duplicates

        for resource in result.get('resources', []):
            public_id = resource['public_id']
            filename = public_id.split('/')[-1]

            # Use etag (file hash) or bytes to identify duplicates
            # Etag is a hash of the actual file content
            file_hash = resource.get('etag') or resource.get('bytes')

            # Skip if we've already seen this exact image content
            if file_hash and file_hash in seen_hashes:
                continue
            if file_hash:
                seen_hashes[file_hash] = filename

            # Generate thumbnail URL (600px max, optimized)
            thumb_url = cloudinary.CloudinaryImage(public_id).build_url(
                width=600,
                height=600,
                crop='limit',
                quality='auto:good',
                fetch_format='auto'
            )

            # Generate full resolution URL (optimized)
            full_url = cloudinary.CloudinaryImage(public_id).build_url(
                quality='auto:best',
                fetch_format='auto'
            )

            photos.append({
                'filename': filename,
                'thumb': thumb_url,
                'full': full_url
            })

        return JsonResponse({'photos': sorted(photos, key=lambda x: x['filename'])}, safe=False)

    except Exception as e:
        print(f"Error fetching from Cloudinary: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'photos': []}, safe=False)


def our_journey(request):
    """Interactive map showing locations from the couple's journey"""
    locations = Location.objects.filter(is_active=True)

    context = {
        'locations': locations,
    }
    return render(request, 'wedding/our_journey.html', context)


def locations_api(request):
    """API endpoint for location data (used by the interactive map)"""
    import cloudinary.api

    locations = Location.objects.filter(is_active=True)

    locations_list = []
    for location in locations:
        # Get photos from Cloudinary for this location
        photo_urls = []

        if location.photo_base_name:
            try:
                # Search for photos matching this location's base name
                result = cloudinary.api.resources(
                    type="upload",
                    prefix=f"wedding/locations/{location.photo_base_name}",
                    max_results=10
                )

                for resource in result.get('resources', []):
                    public_id = resource['public_id']

                    # Progressive JPEG for incremental loading
                    photo_url = cloudinary.CloudinaryImage(public_id).build_url(
                        quality='auto:good',
                        fetch_format='auto',
                        flags='progressive'
                    )

                    photo_urls.append(photo_url)
            except Exception as e:
                print(f"Error fetching Cloudinary photos for {location.location_name}: {e}")

        locations_list.append({
            'id': location.id,
            'location_name': location.location_name,
            'city': location.city,
            'state_country': location.state_country,
            'latitude': float(location.latitude),
            'longitude': float(location.longitude),
            'description': location.description,
            'significance': location.significance,
            'date_visited': location.date_visited,
            'order': location.order,
            'photos': photo_urls,  # Array of Cloudinary photo URLs
        })

    return JsonResponse({'locations': locations_list}, safe=False)


def event_details(request):
    """Wedding event details page"""
    context = {
        'date': 'September 5th, 2026',
        'venue_name': 'The National Aviary',
        'venue_address': 'Pittsburgh, PA',
        'guest_arrival_time': '5:00 PM',
        'ceremony_time': '5:30 PM',
    }
    return render(request, 'wedding/event_details.html', context)


def rsvp(request):
    """RSVP form page"""
    if request.method == 'POST':
        form = RSVPForm(request.POST)
        if form.is_valid():
            rsvp_obj = form.save()

            # Process guest fields from dynamic form
            number_of_guests = rsvp_obj.number_of_guests
            for i in range(number_of_guests - 1):  # -1 because primary person is not a "guest"
                first_name = request.POST.get(f'guest_{i}_first_name', '').strip()
                last_name = request.POST.get(f'guest_{i}_last_name', '').strip()
                use_primary_phone = request.POST.get(f'guest_{i}_use_primary_phone') == 'on'
                guest_phone = request.POST.get(f'guest_{i}_phone', '').strip()

                if first_name and last_name:
                    Guest.objects.create(
                        rsvp=rsvp_obj,
                        first_name=first_name,
                        last_name=last_name,
                        use_primary_phone=use_primary_phone,
                        phone=guest_phone if not use_primary_phone else ''
                    )

            # Send email notification with Excel attachment
            try:
                from django.core.mail import EmailMessage
                from django.conf import settings
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                import io
                from datetime import datetime

                # Create Excel workbook with all RSVPs
                wb = Workbook()
                ws = wb.active
                ws.title = "RSVPs"

                # Add headers with styling - now one row per person
                headers = [
                    'First Name', 'Last Name', 'Contact Phone', 'Email', 'Is Primary Contact',
                    'Attending', 'Dietary Restrictions', 'Song Request', 'Message',
                    'Submitted At', 'Seat Number'
                ]

                header_fill = PatternFill(start_color="C99B8A", end_color="C99B8A", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                # Add all RSVP data - one row per person (primary + guests)
                rsvps = RSVP.objects.all().order_by('-submitted_at')
                current_row = 2

                for rsvp in rsvps:
                    # Add primary contact row
                    ws.cell(row=current_row, column=1, value=rsvp.first_name)
                    ws.cell(row=current_row, column=2, value=rsvp.last_name)
                    ws.cell(row=current_row, column=3, value=rsvp.phone)
                    ws.cell(row=current_row, column=4, value=rsvp.email)
                    ws.cell(row=current_row, column=5, value='Yes')
                    ws.cell(row=current_row, column=6, value=rsvp.get_attendance_display())
                    ws.cell(row=current_row, column=7, value=rsvp.dietary_restrictions)
                    ws.cell(row=current_row, column=8, value=rsvp.song_request)
                    ws.cell(row=current_row, column=9, value=rsvp.message)
                    ws.cell(row=current_row, column=10, value=rsvp.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=current_row, column=11, value='')  # Seat number to be filled later
                    current_row += 1

                    # Add guest rows
                    for guest in rsvp.guests.all():
                        ws.cell(row=current_row, column=1, value=guest.first_name)
                        ws.cell(row=current_row, column=2, value=guest.last_name)
                        ws.cell(row=current_row, column=3, value=guest.get_contact_phone())
                        ws.cell(row=current_row, column=4, value='')  # Guests don't have separate email
                        ws.cell(row=current_row, column=5, value='No')
                        ws.cell(row=current_row, column=6, value=rsvp.get_attendance_display())
                        ws.cell(row=current_row, column=7, value='')  # Individual dietary restrictions not tracked
                        ws.cell(row=current_row, column=8, value='')
                        ws.cell(row=current_row, column=9, value='')
                        ws.cell(row=current_row, column=10, value=rsvp.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
                        ws.cell(row=current_row, column=11, value='')  # Seat number to be filled later
                        current_row += 1

                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

                # Save to bytes
                excel_file = io.BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                # Create email with attachment
                subject = f'New Wedding RSVP: {rsvp_obj.first_name} {rsvp_obj.last_name}'

                # Build guest list
                guest_list = []
                for guest in rsvp_obj.guests.all():
                    guest_list.append(f"  - {guest.first_name} {guest.last_name} (Phone: {guest.get_contact_phone()})")
                guests_text = '\n'.join(guest_list) if guest_list else 'None'

                message = f"""
New RSVP Received!

Primary Contact:
Name: {rsvp_obj.first_name} {rsvp_obj.last_name}
Email: {rsvp_obj.email}
Phone: {rsvp_obj.phone}

RSVP Details:
Attending: {rsvp_obj.get_attendance_display()}
Number of Guests: {rsvp_obj.number_of_guests}

Additional Guests:
{guests_text}

Dietary Restrictions: {rsvp_obj.dietary_restrictions or 'None'}
Song Request: {rsvp_obj.song_request or 'None'}
Message: {rsvp_obj.message or 'None'}

Submitted: {rsvp_obj.submitted_at}

---
See attached Excel file for all RSVPs with complete guest details.
                """

                email = EmailMessage(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    ['loevliedenny@gmail.com', 'caitbmorrow@gmail.com'],
                )

                # Attach Excel file
                filename = f'wedding_rsvps_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                email.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                email.send(fail_silently=False)

            except Exception as e:
                print(f"Error sending email: {e}")
                import traceback
                traceback.print_exc()

            messages.success(request, 'Thank you for your RSVP! We can\'t wait to celebrate with you!')
            return redirect('wedding:rsvp')
    else:
        form = RSVPForm()

    context = {
        'form': form,
    }
    return render(request, 'wedding/rsvp.html', context)


def registry(request):
    """Registry information page"""
    context = {
        'registry_links': [
            {'name': 'With Joy', 'url': 'https://withjoy.com/caitlin-m-and-dennis/registry'},
        ]
    }
    return render(request, 'wedding/registry.html', context)


def photo_upload(request):
    """Photo upload page for guests"""
    from .forms import PhotoUploadForm
    from django.conf import settings
    import cloudinary.uploader

    if request.method == 'POST':
        form = PhotoUploadForm(request.POST, request.FILES)
        if form.is_valid():
            photo_file = request.FILES['photo']

            try:
                # Upload to Cloudinary
                result = cloudinary.uploader.upload(
                    photo_file,
                    folder="wedding/guest_photos",
                    resource_type="image",
                    transformation=[
                        {'quality': 'auto:good'},
                        {'fetch_format': 'auto'}
                    ]
                )

                # Create PhotoUpload object
                photo_upload = form.save(commit=False)
                photo_upload.photo_url = result['secure_url']
                photo_upload.save()

                messages.success(request, 'Thank you for sharing your photo! We can\'t wait to see all the memories.')
                return redirect('wedding:photo_upload')

            except Exception as e:
                print(f"Error uploading photo: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, 'There was an error uploading your photo. Please try again.')
    else:
        form = PhotoUploadForm()

    context = {
        'form': form,
        'compress_photos': getattr(settings, 'COMPRESS_WEDDING_PHOTOS', True),
    }
    return render(request, 'wedding/photo_upload.html', context)


def photo_gallery(request):
    """Photo gallery page displaying all guest uploads"""
    return render(request, 'wedding/photo_gallery.html')


def photos_api(request):
    """API endpoint for photo gallery"""
    photos = PhotoUpload.objects.filter(is_approved=True).order_by('-uploaded_at')

    photos_list = []
    for photo in photos:
        photos_list.append({
            'id': photo.id,
            'uploaded_by': photo.uploaded_by_name,
            'photo_url': photo.photo_url,
            'caption': photo.caption,
            'uploaded_at': photo.uploaded_at.strftime('%B %d, %Y at %I:%M %p'),
        })

    return JsonResponse({'photos': photos_list}, safe=False)


def rsvp_names_api(request):
    """API endpoint to get all RSVP names for autocomplete"""
    names = []

    # Get primary contact names
    rsvps = RSVP.objects.all()
    for rsvp in rsvps:
        names.append(f"{rsvp.first_name} {rsvp.last_name}")

    # Get guest names
    guests = Guest.objects.all()
    for guest in guests:
        names.append(f"{guest.first_name} {guest.last_name}")

    # Remove duplicates and sort
    names = sorted(list(set(names)))

    return JsonResponse({'names': names}, safe=False)


def bachelor_party(request):
    """Bachelor party planning page for NOLA 2026"""
    return render(request, 'wedding/bachelor_party.html')


def things_to_do(request):
    """Curated Pittsburgh food & drink guide for wedding guests"""
    sections = [
        {
            'id': 'food',
            'index': 'I',
            'title': 'Where to Eat',
            'nav_title': 'Food',
            'subtitle': 'Our favorite tables in the city',
            'nav_icon': 'fas fa-utensils',
            'accent': '#c99b8a', 'tint': '#efdcd4',   # dusty rose
            'image': 'https://images.unsplash.com/photo-1750943082452-c714763f73b2?w=1600&h=1000&fit=crop&q=80&fm=jpg',
            'spots': [
                {
                    'name': 'Dish Osteria & Bar', 'fav': True, 'fa': 'fas fa-pizza-slice',
                    'neighborhood': 'South Side', 'tag': 'Sicilian', 'price': '$$$',
                    'signature': 'The "Crack" (their famous rigatoni), or the Spaghetti ai Frutti di Mare — a lighter, spicy seafood pasta',
                    'url': 'https://dishosteria.com/',
                    'maps': 'Dish Osteria 128 S 17th St Pittsburgh PA',
                },
                {
                    'name': 'Café Du Jour', 'fa': 'fas fa-cheese',
                    'neighborhood': 'South Side', 'tag': 'European Bistro', 'price': '$$',
                    'url': 'https://www.cafedujourpgh.com/',
                    'maps': 'Cafe Du Jour 1107 E Carson St Pittsburgh PA',
                },
                {
                    'name': 'Fet-Fisk', 'fa': 'fas fa-fish',
                    'neighborhood': 'Bloomfield', 'tag': 'Nordic Seafood', 'price': '$$$',
                    'signature': 'The Ora King salmon gravlax, or the oysters',
                    'url': 'https://www.fetfisk.net/',
                    'maps': 'Fet-Fisk 4786 Liberty Ave Pittsburgh PA',
                },
                {
                    'name': 'Apteka', 'fa': 'fas fa-seedling',
                    'neighborhood': 'Bloomfield', 'tag': 'Polish Vegan', 'price': '$$',
                    'signature': 'The smażone (fried) pierogi, or the zakąski pickle plate',
                    'url': 'https://aptekapgh.com/',
                    'maps': 'Apteka 4606 Penn Ave Pittsburgh PA',
                },
                {
                    'name': 'Mola', 'fa': 'fas fa-utensils',
                    'neighborhood': 'East Liberty', 'tag': 'Sushi', 'price': '$$$',
                    'signature': 'The blue crab hand roll',
                    'url': 'https://themolafish.com/',
                    'maps': 'Mola 6018 Penn Ave Pittsburgh PA',
                },
            ],
        },
        {
            'id': 'drinks',
            'index': 'II',
            'title': 'Where to Drink',
            'nav_title': 'Drinks',
            'subtitle': 'Cocktails, beer halls & hidden bars',
            'nav_icon': 'fas fa-glass-cheers',
            'accent': '#2c3e50', 'tint': '#cfd6dd',   # slate
            'image': 'https://images.unsplash.com/photo-1752141930096-ac8292d6a15a?w=1600&h=1000&fit=crop&q=80&fm=jpg',
            'spots': [
                {
                    'name': 'Acacia', 'fav': True, 'fa': 'fas fa-glass-martini-alt',
                    'neighborhood': 'South Side', 'tag': 'Speakeasy', 'price': '$$$',
                    'signature': "A bartender's-choice cocktail (tell them your flavors) — or Denny's favorite Aperol spritz",
                    'url': 'https://acaciacocktails.com/',
                    'maps': 'Acacia 2108 E Carson St Pittsburgh PA',
                },
                {
                    'name': 'Lorelei', 'fa': 'fas fa-beer',
                    'neighborhood': 'East Liberty', 'tag': 'Beer Hall & Cocktails', 'price': '$$',
                    'url': 'https://www.loreleipgh.com/',
                    'maps': 'Lorelei 124 S Highland Ave Pittsburgh PA',
                },
                {
                    'name': 'Grapperia', 'fav': True, 'fa': 'fas fa-wine-glass-alt',
                    'neighborhood': 'Lawrenceville', 'tag': 'Grappa & Amaro Bar', 'price': '$$',
                    'signature': 'The "Where\'s Tony?" cocktail',
                    'url': 'http://www.grapperiapgh.com/',
                    'maps': 'Grapperia 3801 Butler St Pittsburgh PA',
                },
                {
                    'name': 'Hidden Harbor', 'fa': 'fas fa-cocktail',
                    'neighborhood': 'Squirrel Hill', 'tag': 'Tiki Bar', 'price': '$$$',
                    'signature': 'The Rum Bucket',
                    'url': 'https://www.hiddenharborpgh.com/',
                    'maps': 'Hidden Harbor Squirrel Hill Pittsburgh PA',
                },
            ],
        },
        {
            'id': 'breweries',
            'index': 'III',
            'title': 'Local Breweries',
            'nav_title': 'Breweries',
            'subtitle': 'For the beer lovers',
            'nav_icon': 'fas fa-beer',
            'accent': '#c9a227', 'tint': '#eaddb0',   # gold (slightly deepened for contrast)
            'image': 'https://images.unsplash.com/photo-1663060435381-7a3d91a59836?w=1600&h=1000&fit=crop&q=80&fm=jpg',
            'spots': [
                {
                    'name': 'Eleventh Hour Brewing Co.', 'fa': 'fas fa-beer',
                    'neighborhood': 'Lawrenceville', 'tag': 'Brewery', 'price': '$$',
                    'signature': "Any of the sours — and you're right next door to Grapperia",
                    'url': 'https://www.11thhourbrews.com/',
                    'maps': 'Eleventh Hour Brewing 3711 Charlotte St Pittsburgh PA',
                },
                {
                    'name': 'Dancing Gnome', 'fa': 'fas fa-beer',
                    'neighborhood': 'Millvale', 'tag': 'Brewery', 'price': '$$',
                    'signature': 'Lustra, the flagship Citra-and-Amarillo pale ale',
                    'url': 'https://dancinggnomebeer.com/',
                    'maps': 'Dancing Gnome Brewery Millvale PA',
                },
            ],
        },
        {
            'id': 'sweets',
            'index': 'IV',
            'title': 'Coffee & Something Sweet',
            'nav_title': 'Sweets',
            'subtitle': 'A little something extra',
            'nav_icon': 'fas fa-mug-hot',
            'accent': '#8ba17f', 'tint': '#d3ddcb',   # sage (deepened for contrast)
            'image': 'https://images.unsplash.com/photo-1553962311-62f2471b159d?w=1600&h=1000&fit=crop&q=80&fm=jpg',
            'spots': [
                {
                    'name': "Delanie's Coffee", 'fa': 'fas fa-mug-hot',
                    'neighborhood': 'South Side', 'tag': 'Coffee', 'price': '$',
                    'signature': "A drip of Partners beans — what Denny still brews at home today",
                    'url': 'https://www.delaniescoffee.com/',
                    'maps': "Delanie's Coffee 1737 E Carson St Pittsburgh PA",
                },
                {
                    'name': "Page's Dairy Mart", 'fav': True, 'fa': 'fas fa-ice-cream',
                    'neighborhood': 'South Side', 'tag': 'Ice Cream', 'price': '$',
                    'signature': 'The pumpkin ice cream, if they have it — a must',
                    'url': 'https://www.pagesdairymart.com/',
                    'maps': "Page's Dairy Mart 4112 E Carson St Pittsburgh PA",
                    'seasonal': 'Open seasonally, roughly Memorial Day–Labor Day',
                },
            ],
        },
        {
            'id': 'explore',
            'index': 'V',
            'title': 'What to See & Do',
            'nav_title': 'Explore',
            'subtitle': 'Iconic views, gardens, museums & neighborhoods',
            'nav_icon': 'fas fa-camera-retro',
            'accent': '#a9714e', 'tint': '#e4cdb8',   # terracotta (the incline's red brick)
            'image': 'https://upload.wikimedia.org/wikipedia/commons/thumb/4/43/Strip_District_at_Dusk%2C_October_2019.jpg/1920px-Strip_District_at_Dusk%2C_October_2019.jpg',
            'mustry_label': "Don't miss",
            'spots': [
                {
                    'name': 'Duquesne Incline', 'fa': 'fas fa-mountain',
                    'neighborhood': 'Mount Washington', 'tag': 'Historic Funicular',
                    'url': 'https://www.duquesneincline.org/',
                    'maps': 'Duquesne Incline 1197 W Carson St Pittsburgh PA',
                },
                {
                    'name': 'The Strip District', 'fav': True, 'fa': 'fas fa-store',
                    'neighborhood': 'Strip District', 'tag': 'Market Neighborhood',
                    'maps': 'Strip District Penn Ave Pittsburgh PA',
                },
                {
                    'name': 'Pittsburgh Zoo & Aquarium', 'fa': 'fas fa-paw',
                    'neighborhood': 'Highland Park', 'tag': 'Zoo & Aquarium',
                    'url': 'https://www.pittsburghzoo.org/',
                    'maps': 'Pittsburgh Zoo and Aquarium 7370 Baker St Pittsburgh PA',
                },
                {
                    'name': 'Soldiers & Sailors Memorial Hall & Museum', 'fa': 'fas fa-monument',
                    'neighborhood': 'Oakland', 'tag': 'Historic Landmark & Museum',
                    'url': 'https://www.soldiersandsailorshall.org/',
                    'maps': 'Soldiers and Sailors Memorial Hall 4141 Fifth Ave Pittsburgh PA',
                },
                {
                    'name': 'Phipps Conservatory and Botanical Gardens', 'fa': 'fas fa-spa',
                    'neighborhood': 'Oakland', 'tag': 'Gardens & Conservatory',
                    'url': 'https://www.phipps.conservatory.org/',
                    'maps': 'Phipps Conservatory One Schenley Dr Pittsburgh PA',
                },
                {
                    'name': 'The Andy Warhol Museum', 'fa': 'fas fa-palette',
                    'neighborhood': 'North Shore', 'tag': 'Modern Art Museum',
                    'url': 'https://www.warhol.org/',
                    'maps': 'The Andy Warhol Museum 117 Sandusky St Pittsburgh PA',
                },
            ],
        },
    ]

    # Continuous editorial numbering (01, 02, ...) across all sections
    counter = 1
    for section in sections:
        for spot in section['spots']:
            spot['num'] = '{:02d}'.format(counter)
            counter += 1

    return render(request, 'wedding/things_to_do.html', {'sections': sections})
