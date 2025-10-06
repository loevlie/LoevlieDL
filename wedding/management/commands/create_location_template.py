"""
Management command to create a sample XLSX template for locations.

Usage:
    python manage.py create_location_template [output_file.xlsx]
"""

from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill


class Command(BaseCommand):
    help = 'Create a sample XLSX template for importing locations'

    def add_arguments(self, parser):
        parser.add_argument(
            'output_file',
            nargs='?',
            type=str,
            default='location_template.xlsx',
            help='Output file path (default: location_template.xlsx)'
        )

    def handle(self, *args, **options):
        output_file = options['output_file']

        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Locations"

            # Define headers
            headers = [
                'location_name',
                'city',
                'state_country',
                'latitude',
                'longitude',
                'description',
                'date_visited',
                'significance',
                'order',
                'is_active'
            ]

            # Style for headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            # Write headers
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            # Add sample data
            sample_data = [
                {
                    'location_name': 'West Virginia University',
                    'city': 'Morgantown',
                    'state_country': 'West Virginia',
                    'latitude': 39.6498,
                    'longitude': -79.9545,
                    'description': 'Where we met as classmates and fell in love.',
                    'date_visited': '2018-2022',
                    'significance': 'The place where our story began',
                    'order': 1,
                    'is_active': True
                },
                {
                    'location_name': 'The Aviary',
                    'city': 'Pittsburgh',
                    'state_country': 'Pennsylvania',
                    'latitude': 40.4406,
                    'longitude': -79.9959,
                    'description': 'Our home city where we built our life together.',
                    'date_visited': '2022-Present',
                    'significance': 'Where we chose to build our future',
                    'order': 2,
                    'is_active': True
                },
            ]

            # Write sample data
            for row_num, data in enumerate(sample_data, start=2):
                for col_num, header in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data.get(header, ''))

            # Adjust column widths
            column_widths = {
                'A': 30,  # location_name
                'B': 20,  # city
                'C': 20,  # state_country
                'D': 12,  # latitude
                'E': 12,  # longitude
                'F': 50,  # description
                'G': 20,  # date_visited
                'H': 40,  # significance
                'I': 10,  # order
                'J': 10,  # is_active
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Save workbook
            workbook.save(output_file)

            self.stdout.write(self.style.SUCCESS(f'✓ Template created: {output_file}'))
            self.stdout.write(self.style.SUCCESS('\nColumns:'))
            for header in headers:
                self.stdout.write(f'  - {header}')
            self.stdout.write(self.style.SUCCESS('\nYou can now edit this file and import it using:'))
            self.stdout.write(self.style.SUCCESS(f'  python manage.py import_locations {output_file}'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Failed to create template: {str(e)}'))
